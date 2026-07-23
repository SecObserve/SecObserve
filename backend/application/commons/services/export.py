import json
import logging
from datetime import datetime
from typing import Any, Optional

import jsonpickle
from defusedcsv import csv
from django.db.models.query import QuerySet
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font

logger = logging.getLogger("secobserve.commons")


def _escape_formula(value: Any) -> Any:
    if not value or not isinstance(value, str):
        return value

    # Removing illegal XML characters: Excel cannot handle certain control
    # characters (such as ASCII 0-31)
    cleaned = ILLEGAL_CHARACTERS_RE.sub("", value)

    # Neutralize spreadsheet formula injection in string cells, matching what
    # defusedcsv does for CSV: prefix a leading formula trigger with a quote.
    if cleaned[0] in ("=", "+", "-", "@", "\t", "\r"):
        cleaned = "'" + cleaned

    return cleaned


def export_excel(objects: QuerySet, title: str, excludes: list[str], foreign_keys: list[str]) -> Workbook:
    workbook = Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active
    worksheet.title = title

    font_bold = Font(bold=True)

    row_num = 1

    for current_object in objects:
        if row_num == 1:
            col_num = 1
            for key in dir(current_object):
                if key not in excludes and not callable(getattr(current_object, key)) and not key.startswith("_"):
                    value = key.replace("_", " ").capitalize()
                    cell = worksheet.cell(row=row_num, column=col_num, value=value)
                    cell.font = font_bold
                    col_num += 1
            cell.font = font_bold
            row_num = 2
        if row_num > 1:
            col_num = 1
            for key in dir(current_object):
                if key not in excludes and not callable(getattr(current_object, key)) and not key.startswith("_"):
                    value = current_object.__dict__.get(key)
                    if key in foreign_keys and getattr(current_object, key):
                        value = str(getattr(current_object, key))
                    if value and isinstance(value, datetime):
                        value = value.replace(tzinfo=None)
                    if value and isinstance(value, (dict, list)):
                        value = str(value)
                    value = _escape_formula(value)
                    try:
                        worksheet.cell(row=row_num, column=col_num, value=value)
                    except Exception as e:
                        logger.warning("Cannot set cell with type %s", type(value))
                        logger.warning(str(e))
                    col_num += 1
        row_num += 1

    return workbook


def export_csv(
    response: HttpResponse,
    objects: QuerySet,
    excludes: list[str],
    foreign_keys: list[str],
) -> None:
    writer = csv.writer(response)  # nosemgrep
    # defusedcsv is actually used but not detected by Semgrep

    first_row = True

    for current_object in objects:
        fields: list[Any] = []
        if first_row:
            fields.clear()
            for key in dir(current_object):
                if key not in excludes and not callable(getattr(current_object, key)) and not key.startswith("_"):
                    value = key.replace("_", " ").capitalize()
                    fields.append(value)

            writer.writerow(fields)

            first_row = False
        if not first_row:
            fields.clear()
            for key in dir(current_object):
                if key not in excludes and not callable(getattr(current_object, key)) and not key.startswith("_"):
                    value = current_object.__dict__.get(key)
                    if key in foreign_keys and getattr(current_object, key):
                        value = str(getattr(current_object, key))
                    if value and isinstance(value, str):
                        value = value.replace("\n", " NEWLINE ").replace("\r", "")
                    fields.append(value)

            writer.writerow(fields)


def object_to_json(object_to_encode: Any) -> str:
    jsonpickle.set_encoder_options("json", ensure_ascii=False)
    json_string = jsonpickle.encode(object_to_encode, unpicklable=False)

    json_dict = json.loads(json_string)
    json_dict = _remove_empty_elements(json_dict)

    return json.dumps(json_dict, indent=4, sort_keys=True, ensure_ascii=False)


def _remove_empty_elements(d: dict | list) -> dict | list:
    """recursively remove empty lists, empty dicts, or None elements from a dictionary"""

    def empty(x: Optional[(dict | list)]) -> bool:
        return x is None or x == {} or x == []

    if not isinstance(d, (dict | list)):
        return d
    if isinstance(d, list):
        return [v for v in (_remove_empty_elements(v) for v in d) if not empty(v)]

    return {k: v for k, v in ((k, _remove_empty_elements(v)) for k, v in d.items()) if not empty(v)}
